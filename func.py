# import xlrd, xlwings
from collections import namedtuple
from fuzzywuzzy import fuzz
from fuzzywuzzy import process as fuzz_process
from colorama import Fore, Style, Back
import os.path
import math
import pandas as pd
import numpy as np
from datetime import date
from warnings import filterwarnings
import datetime
import time
import traceback
import sys
import settings
import pyinputplus as pyip

filterwarnings('ignore', category=UserWarning, module='openpyxl')


# @numerated_list
# def list_files_in_folder(folder_path_to_list: str) -> list:
#     # todo drop function if possible
#     files_list = check_excel_files_list(folder_path_to_list)
#     return files_list
#
#
# def select_files_to_load(folder_path_to_list, sheet_name=None):
#     file_name_list_to_load = []
#     try:
#         file_list = list_files_in_folder(folder_path_to_list)
#
#     except FileNotFoundError:
#         print(Fore.RED + 'WARNING: Specified directory has not been found' + Style.RESET_ALL)
#         sys.exit()
#
#     # Input file number from displayed list and load Excel file
#     hint = 'Which file do you want to download (input a number above): '
#     limit = [i for i in range(len(file_list))]
#     file_number = user_input(input_type=int, limit=limit, label=hint, distinct=True)
#
#     df: pd.DataFrame = pd.read_excel(os.path.join(folder_path_to_list, file_list[file_number - 1]),
#                                      sheet_name=sheet_name)
#     df: pd.DataFrame = pd.ExcelFile(d)
#
#     return df


class Folder:

    @staticmethod
    def __isdir_check(x):
        check = os.path.isdir(x)
        if not check:
            raise IsADirectoryError
        else:
            return x

    def __init__(self, folder_path: str):
        self.path = Folder.__isdir_check(folder_path)
        self.name = os.path.basename(folder_path)

    @property
    def files(self) -> list | str:

        wrong = ['~$']
        file_list = []

        try:
            for file in os.listdir(self.path):
                for bad in wrong:
                    if not os.path.isdir(os.path.join(self.path, file)) and bad not in file:
                        file_list.append(file)

            return file_list

        except FileNotFoundError:
            print(Fore.RED + 'Specified directory has not been found' + Style.RESET_ALL)

    @property
    def excel_files(self):

        excel_extension = '.xls'
        files_list = [file for file in self.files if excel_extension in file]
        return files_list

    def select_file(self, xls=True):
        file_name = pyip.inputMenu(choices=self.excel_files if xls else self.files,
                                   prompt='Files:\n',
                                   numbered=True,
                                   blank=True)

        print('Selected:', Fore.GREEN + file_name + Style.RESET_ALL)
        return file_name

    def file_exists(self, file: str):
        return os.path.exists(os.path.join(self.path, file))


def timer(some_process):
    def wrapper(*args, **kwargs):
        time_started = datetime.datetime.now()
        some_process(*args, **kwargs)
        time_finished = datetime.datetime.now()
        process_time = time_finished - time_started  # TODO add time format
        print('(', process_time, ' sec)', sep='')

    return wrapper


def process_decorator(process_function):
    @timer
    def wrapper(*args, **kwargs):
        print('--------------')
        print('Process has just started')
        print('--------------')
        process_function(*args, **kwargs)
        print(Fore.GREEN + 'Finished successfully' + Style.RESET_ALL, end=' ')

    return wrapper


class MatchingProcess:
    # TODO matching process as service class

    def __init__(self, choices: pd.DataFrame | pd.Series | list):
        self.choices = choices
        self.result: pd.DataFrame = pd.DataFrame()

    def best_sequence_match(self, query: pd.Series | list,
                            score_cutoff_choice=80,
                            show_results_in_terminal=False,
                            drop_not_best: bool = True,
                            add_column_name: str | bool = False
                            ) -> pd.DataFrame:

        for series_name in self.choices:
            mapping_df = self.choices
            self.choices = self.choices[series_name]

            res = self.sequence_match(query=query,
                                      score_cutoff_choice=score_cutoff_choice,
                                      show_results_in_terminal=show_results_in_terminal)

            self.result = pd.concat([self.result, res], axis=0)

            self.choices = mapping_df

        self.result.reset_index(drop=True, inplace=True)

        if drop_not_best:
            ind_max = self.result.groupby(['query'])['score'].idxmax()
            self.result = self.result.loc[ind_max, :]

        self.result.index = self.result['result_index']
        self.result.drop(columns='result_index', inplace=True)

        if add_column_name:
            self.result['column'] = add_column_name

        return self.result

    def single_match(self,
                     query: str,
                     score_cutoff_choice: int = 80,
                     show_results_in_terminal: bool = False,
                     show_matched: bool = True) -> str | list:

        scorers_coef = {fuzz.token_sort_ratio: 0.11,  # 1
                        fuzz.QRatio: 0.11,  # 2
                        fuzz.UQRatio: 0.11,  # 3
                        fuzz.UWRatio: 0.11,  # 4
                        fuzz.WRatio: 0.11,  # 5
                        fuzz.partial_ratio: 0.11,  # 6
                        fuzz.ratio: 0.11,  # 7
                        fuzz.partial_token_set_ratio: 0.11,  # 8
                        fuzz.partial_token_sort_ratio: 0.11  # 9
                        }

        # check query str

        if isinstance(self.choices, pd.Series):
            choices = self.choices.dropna().tolist()
        else:
            choices = self.choices

        # processors = []   integrate processors
        results = {x: 0 for x in choices}

        for scorer in scorers_coef:
            # for processor in processors:
            match = fuzz_process.extractOne(query=query,
                                            choices=self.choices,
                                            scorer=scorer,
                                            # processor=processor,
                                            score_cutoff=score_cutoff_choice
                                            )
            if match is None:
                continue

            results[match[0]] += match[1] * scorers_coef[scorer]

        match_process_result = (MatchingProcess
                                .check_mapping_results(results,
                                                       query,
                                                       score_cutoff_choice,
                                                       show_results_in_terminal
                                                       )
                                )
        # add index in default table

        if isinstance(self.choices, pd.Series):
            result_index = self.choices[self.choices == match_process_result[1]].index.tolist()[0]
            match_process_result.append(result_index)

        if show_matched:
            return match_process_result
        else:
            return match_process_result[1]

    @staticmethod
    def check_mapping_results(matching_results: dict,
                              query: str,
                              score_cutoff_choice: int,
                              show_results_in_terminal: bool
                              ) -> list:

        max_score_match = max(matching_results, key=matching_results.get)
        max_score = max(matching_results.values())
        label_matched = 'matched'
        label_did_not_matched = 'not found'

        label = label_did_not_matched if max_score < score_cutoff_choice else label_matched

        if show_results_in_terminal:
            print('----------------')
            print(query, ' -- ', max_score_match, f' ({label})')
            print('----------------')
            print(matching_results)
            print()
        return [query, max_score_match, label, max_score]

    def sequence_match(self,
                       query: list | pd.Series,
                       score_cutoff_choice: int = 75,
                       show_results_in_terminal: bool = False
                       ) -> pd.DataFrame:

        result = dict()

        # Get exact structure
        if isinstance(query, pd.Series):
            input_list_process = query.dropna().unique()
        else:
            input_list_process = set([x for x in query if str(x) != 'nan'])

        for name in input_list_process:
            processed_choice = self.single_match(query=name,
                                                 score_cutoff_choice=score_cutoff_choice,
                                                 show_results_in_terminal=show_results_in_terminal
                                                 )

            result[name] = processed_choice

        return MatchingProcess.parse_match_results(result)

    @staticmethod
    def parse_match_results(match_results: dict) -> pd.DataFrame:
        dataframe_columns = ['query', 'result', 'matched', 'score', 'result_index']
        data = []

        for value in match_results.values():
            data.append(value)

        match_results_dataframe = pd.DataFrame(data, columns=dataframe_columns)

        return match_results_dataframe

    def mapping_table(self,
                      query: pd.Series):

        # self.choice must be pd.DataFrame, not pd.Series
        result_names = self.choices.iloc[:, 0]
        self.choices = self.choices.iloc[:, 1:]

        self.best_sequence_match(query=query, drop_not_best=True, add_column_name=query.name)
        self.result.loc[:, 'chosen'] = result_names
        self.result.reset_index(drop=True, inplace=True)

        return self.result


class BasicProcessClass:

    def __init__(self, data: dict | pd.DataFrame, process_name: str = 'untitled'):
        self.name = process_name
        self.data = data
        self.user_choice: int | None = None
        self.result = None

    @process_decorator
    def run(self):
        pass

    def save_file(self, dataframes, folder_path, file_name, ):
        return self


def single_input_file(folder_path):
    list_of_folder_excel_files = check_excel_files_list(folder_path)
    file_number_to_load = pyip.inputNum(min=1, max=len(list_of_folder_excel_files)) - 1
    file_name_to_load = list_of_folder_excel_files[file_number_to_load]

    return file_name_to_load


def numerated_list(func):
    def wrapper(arg):
        folder_list = func(arg)
        print()
        print(f'Files in "{arg}" displayed below')
        print('--------------')
        for ind, file in enumerate(folder_list):
            print(ind + 1, '. ', file, sep='')
        print('--------------')
        print('Choose one file (input a number)')

        return folder_list

    return wrapper


@numerated_list
def check_excel_files_list(folder_path: str) -> list:
    extensions = ['.xlsx', '.xls', '.xlsm', '.xlsb']
    wrong = ['~$']
    correct_list = []

    file_list = [file for file in os.listdir(folder_path)
                 if not os.path.isdir(os.path.join(folder_path, file))]
    file_list_lower = [file.lower() for file in file_list]

    for file_lower, file in zip(file_list_lower, file_list):
        for ext in extensions:
            for wr in wrong:
                if ext in file_lower and wr not in file_lower:
                    correct_list.append(file)

    return list(set(correct_list))


def find_children(file_path):
    parent_dict = dict()
    parent_name = os.path.basename(os.path.normpath(file_path))
    parent_dict[parent_name] = list()
    children = os.listdir(file_path)

    for child in children:

        if os.path.isdir(os.path.join(file_path, child)):
            child_path = os.path.join(file_path, child)
            child_hierarchy = find_children(child_path)
            parent_dict[parent_name].append(child_hierarchy)
        else:
            parent_dict[parent_name].append(child)

    return parent_dict


def get_file_size(file_path):
    bytes_in_megabyte = 1024 * 1024
    size = os.path.getsize(file_path) / bytes_in_megabyte

    return size


def get_create_date(file_path):
    # s1 = path.getctime(file_path)
    s1 = os.stat(file_path).st_birthtime
    s2 = datetime.datetime.strptime(time.ctime(s1), '%c').date()
    s3 = s2.strftime('%d/%m/%Y')

    return s3


def get_mod_date(file_path):
    s1 = os.path.getmtime(file_path)
    s2 = datetime.datetime.strptime(time.ctime(s1), '%c').date()
    s3 = s2.strftime('%d/%m/%Y')

    return s3


def tmp(df: pd.DataFrame):
    df = df.applymap(lambda x: x.strip() if isinstance(x, str) else x)
    return df


def add_extracting_time_info(df: pd.DataFrame):
    extract_timestamp = datetime.datetime.now()

    df['extracting_date'] = extract_timestamp.strftime('%d.%m.%Y')
    df['extracting_time'] = extract_timestamp.strftime('%H:%M')

    return df


def add_edit_time_info(df: pd.DataFrame, fem_folder, filename):
    edit_timestamp = os.path.getmtime(os.path.join(fem_folder, filename))
    datestamp = datetime.datetime.fromtimestamp(edit_timestamp)

    df['edit_date'] = datestamp.strftime('%d.%m.%Y')
    df['edit_time'] = datestamp.strftime('%H:%M')

    return df


def check_duplicate_column_name(df: pd.DataFrame):
    duplicates = [column_name for column_name in df.columns.to_list() if df.columns.to_list().count(column_name) > 1]
    return duplicates


def safe_dataframes_to_excel(dataframes: list,
                             sheet_names: list,
                             folder_to_save=os.path.join(settings.base_location, 'tmp'),
                             file_name='untitled'):
    print('--------------')
    print('File saving')
    print('--------------')

    today_date = date.today().strftime("%d%m%Y")

    def save_process(exist=True):

        path_to_save = f'{folder_to_save}/{file_name}_{today_date}.xlsx'

        if os.path.exists(path_to_save) and exist:
            raise FileExistsError

        with pd.ExcelWriter(path_to_save) as writer:
            for df, sheet_name in zip(dataframes, sheet_names):
                df.to_excel(writer, sheet_name=sheet_name, index=False)
        print(Fore.GREEN + f'File "{file_name}" has been saved to "{folder_to_save}"' + Style.RESET_ALL)

    try:
        save_process()

    except FileNotFoundError:
        os.mkdir(folder_to_save)
        save_process()
    except FileExistsError:
        print(Fore.YELLOW + f'"{file_name}_{today_date}" exists in "{folder_to_save}"' + Style.RESET_ALL)
        print('Do you want to rename file?')
        choice = pyip.inputYesNo(prompt='(y/n) - ', yesVal='y', noVal='n')
        if choice == 'y':
            file_name = input('file_name: ')
            save_process()
        else:
            save_process(exist=False)


@numerated_list
def list_files_in_folder(folder_path_to_list: str) -> list:
    # todo drop function if possible
    files_list = check_excel_files_list(folder_path_to_list)
    return files_list


def select_files_to_load(folder_path_to_list, sheet_name=None):
    file_name_list_to_load = []
    try:
        file_list = list_files_in_folder(folder_path_to_list)

    except FileNotFoundError:
        print(Fore.RED + 'WARNING: Specified directory has not been found' + Style.RESET_ALL)
        sys.exit()

    # Input file number from displayed list and load Excel file
    hint = 'Which file do you want to download (input a number above): '
    limit = [i for i in range(len(file_list))]
    file_number = user_input(input_type=int, limit=limit, label=hint, distinct=True)

    df: pd.DataFrame = pd.read_excel(os.path.join(folder_path_to_list, file_list[file_number - 1]),
                                     sheet_name=sheet_name)
    df: pd.DataFrame = pd.ExcelFile(d)

    return df


def groupby_key_field(df: pd.DataFrame, key: list, column_name='value', aggr_func='sum'):
    # Incorrect function

    df = df.set_index(key).groupby(by=key).sum()
    df = df.reset_index()

    return df


def rename_dict(a_list: list, dictionary: dict, minscore: int):
    # if column has not been found or score equals then store name to mapping
    # todo get rid!
    score = 0
    zeros = 0
    tmp_dict = {}
    dict_for_rename = {}

    # cols_default = [col for col in column_name.unique().tolist() if str(col) != 'nan']
    cols_default = [col for col in set(a_list) if str(col) != 'nan']
    cols_lower = [col.lower() for col in cols_default]
    for col_1, col_2 in zip(cols_default, cols_lower):
        for key, val in dictionary.items():
            for part in val:
                # if col_2.__contains__(part):
                if part in col_1:
                    score += 1
                if part in col_2:
                    score += 1
            tmp_dict[key] = score
            score = 0

        for value in tmp_dict.values():
            if value != 0:
                zeros += 1

        if zeros == 0 or max(tmp_dict.values()) < minscore:
            max_score_name = 'not found'
        else:
            max_score_name = max(tmp_dict, key=tmp_dict.get)

        dict_for_rename[col_1] = max_score_name
        zeros = 0
        tmp_dict = {}

    return dict_for_rename


def read_multiple_excel_files(files_path_list: list, sheet_list: list):
    # todo move to Folder class
    dataframe_dictionary = dict()

    for file, sheet_name in zip(files_path_list, sheet_list):

        file_base_name = os.path.basename(file)
        try:
            load = pd.ExcelFile(file).parse(sheet_name=sheet_name)
            dataframe_dictionary[file_base_name] = load
        except:
            dataframe_dictionary[file] = None
            print(f'{file_base_name} has not been loaded')
            continue

    return dataframe_dictionary


def find_the_best_match():
    # todo move to MatchingProcess class
    # Function loops through columns in mapping file and finds the best match
    pass


def match_process(query: int,
                  choices: list,
                  score_cutoff_choice=80,
                  show_results_in_terminal=False,
                  show_matched=False) -> str | list:
    scorers_coef = {fuzz.token_sort_ratio: 0.11,  # 1
                    fuzz.QRatio: 0.11,  # 2
                    fuzz.UQRatio: 0.11,  # 3
                    fuzz.UWRatio: 0.11,  # 4
                    fuzz.WRatio: 0.11,  # 5
                    fuzz.partial_ratio: 0.11,  # 6
                    fuzz.ratio: 0.11,  # 7
                    fuzz.partial_token_set_ratio: 0.11,  # 8
                    fuzz.partial_token_sort_ratio: 0.11  # 9
                    }

    # processors = []   integrate processors
    results = {x: 0 for x in choices}

    for scorer in scorers_coef:
        # for processor in processors:
        match = fuzz_process.extractOne(query=query,
                                        choices=choices,
                                        scorer=scorer,
                                        # processor=processor,
                                        score_cutoff=score_cutoff_choice
                                        )
        if match is None:
            continue

        results[match[0]] += match[1] * scorers_coef[scorer]

    match_process_result = check_mapping_results(results,
                                                 query,
                                                 score_cutoff_choice,
                                                 show_results_in_terminal
                                                 )

    if show_matched:
        return match_process_result
    else:
        return match_process_result[1]


def check_mapping_results(matching_results: dict,
                          query: str,
                          score_cutoff_choice: int,
                          show_results_in_terminal: bool) -> list:
    # todo move to MatchingProcess class
    max_score_match = max(matching_results, key=matching_results.get)
    max_score = max(matching_results.values())
    label_matched = 'matched'
    label_did_not_matched = 'not found'

    label = label_did_not_matched if max_score < score_cutoff_choice else label_matched

    if show_results_in_terminal:
        print('----------------')
        print(query, ' -- ', max_score_match, ' (matched)')
        print('----------------')
        print(matching_results)
        print()
    return [query, max_score_match, label, max_score]


def fuzzmatch(input_list: list | pd.Series,
              choices: list | pd.Series,
              score_cutoff_choice: int = 75,
              show_results_in_terminal: bool = False,
              show_matched: bool = False) -> list | pd.Series | dict:
    # todo move to MatchingProcess class
    match_results = dict()

    # Get exact structure
    if isinstance(input_list, pd.Series):
        input_list_process = input_list.unique().tolist()
    else:
        input_list_process = input_list

    if isinstance(choices, pd.Series):
        choices = choices.unique().tolist()
    else:
        choices = choices

    for name in input_list_process:
        processed_choice = match_process(query=name,
                                         choices=choices,
                                         score_cutoff_choice=score_cutoff_choice,
                                         show_results_in_terminal=show_results_in_terminal,
                                         show_matched=show_matched)
        match_results[name] = processed_choice

    # if isinstance(input_list, pd.Series):
    #     result = input_list.map(match_results)
    # else:
    #     result = match_results.values()

    # Add more complex logic

    # return result
    return match_results


def parse_match_results(match_results: dict) -> pd.DataFrame:
    # todo move to MatchingProcess class
    dataframe_columns = ['query', 'result', 'matched', 'score']
    data = []

    for value in match_results.values():
        data.append(value)

    match_results_dataframe = pd.DataFrame(data, columns=dataframe_columns)

    return match_results_dataframe


def save_formulas_to_excel():
    import openpyxl
    from openpyxl.cell import Cell
    from openpyxl import Workbook
    wb: Workbook = openpyxl.load_workbook(settings.OPENPYXL_EXTRACTION_PATH)
    ws = wb[settings.openpyxl_sheet_name]
    cellObj: Cell
    for i, cellObj in enumerate(ws['C'], 1):
        if i == 1:
            continue
        print(ws['B'][0])

        cellObj.value = "=vlookup(A{0}, A:B, 2, 0)".format(i)
    wb.save(settings.OPENPYXL_EXTRACTION_PATH)


def check_function():
    pass


if __name__ == '__main__':
    check_function()
    print()
